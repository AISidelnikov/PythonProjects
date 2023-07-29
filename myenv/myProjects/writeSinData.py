from openpyxl import Workbook
import numpy as np
import click

@click.command()
@click.option("-f", default=1, help="The frequency of the signal")
@click.option("-fs", default=1, help="Sample rate")
def create_data_sin(f, fs):
    x = np.arange(fs)
    y = np.sin(2*np.pi * f * (x/fs))

    wb = Workbook()
    ws = wb.active
    ws.title = "sin"

    ws['A1'] = 'x'
    ws['B1'] = 'y'

    for i, j in enumerate(x):
        ws.cell(row=i+2, column=1, value=j)

    for i, j in enumerate(y):
        ws.cell(row=i+2, column=2, value=j)

    wb.save('sin.xlsx')

if __name__ == "__main__":
    create_data_sin()